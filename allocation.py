import pyodbc
import pandas as pd
import os
import openpyxl as pyxl
import xlrd
import datetime

 

today_date = datetime.date.today()

today_datetime = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

############################## Stating Path Variables #########################

 

working_directory = 'C:\\Users\\jacoen\\Desktop\\Allocation'

"""

    The following files need to be saved in the above working directory:

        Main.py

        Allocator1.py

        Allocator2.py

        Allocation sheet.xlsx

        JsonFiles folder

        AllocationResults folder

    Remeber to update working_directory in Allocator1.py and in Allocator2.py

"""

 

 

###############################################################################

 

#Set working directory

os.chdir(working_directory)
 

wb = pyxl.load_workbook('Allocation sheet.xlsx')


def _create_db_connection(**kwargs):

    """

    Establish a connection to the DB

    :param kwargs: db_connection(server name or IP)

    :param kwargs: database

    :param kwargs: username

    :param kwargs: password

    :return: The DB connection object

    """

    driver_type = '{ODBC Driver 13 for SQL Server}'

    connection_str = f'DRIVER={driver_type};' \

                     f'SERVER={kwargs.get("db_connection")};' \

                     f'DATABASE={kwargs.get("database")};UID={kwargs.get("username")};' \

                     f'PWD={kwargs.get("password")}'

    try:

        connection = pyodbc.connect(connection_str)

    except pyodbc.OperationalError as error_msg:

        msg = f'Verify that the server value is correct: {kwargs.get("db_connection")}'

        raise Exception(f'{msg}\n{error_msg.args}')

    return connection

 

 

def _run_db_query(**kwargs):

    """

    Run any query

    :param kwargs: db_connection

    :param kwargs: database

    :param kwargs: username

    :param kwargs: password

    :param kwargs: query

    :return: a tuple of the query results connection object

    """

    results = []

    connection = _create_db_connection(**kwargs)

 

    with connection:

        cursor = connection.cursor()

    query = kwargs.get('query')

    if kwargs.get("db_update"):

        query = kwargs.get('update_query')

    rows, connection = cursor.execute(query), connection

 

    for row in rows:

        results.append(row)

    return results

   

 

kwargs = {'db_connection': 'sql-siebel.int.gray.net,64495',

                           'username': 'sadmin',

                          'password': 'sadminint',

                          'database': 'siebeldb'}

 

query = {"query": "select s.SR_NUM,\

         c.PER_TITLE_SUFFIX,\

         c.FST_NAME,\

         c.LAST_NAME,\

         o.DUNS_NUM,\

         o.NAME,\

         s.CREATED,\

         pl.NAME,\

         sx.ATTRIB_39,\

         s.INS_PRODUCT,\

         s.SR_AREA,\

         s.SR_STAT_ID,\

         s.X_ADMINISTRATOR,\

         u2.LOGIN\

         from S_SRV_REQ s with (nolock) \

         left outer join S_SRV_REQ_X sx with (nolock) on s.ROW_ID = sx.PAR_ROW_ID\

         left outer join S_PROD_LN pl with (nolock) on sx.ATTRIB_46 = pl.ROW_ID \

         left outer join S_ORG_EXT o with (nolock) on o.ROW_ID = s.CST_OU_ID \

         left outer join S_CONTACT c with (nolock) on c.ROW_ID = s.CST_CON_ID\

         left outer join S_USER u2 with (nolock) on s.OWNER_EMP_ID = u2.PAR_ROW_ID \

         where u2.LOGIN = 'UT_ADMIN_TEST'"}

         #and s.CREATED > '15 Jul 2018'"}

         #and s.CREATED < concat(convert(varchar(10), GETDATE(), 120),' ',convert(varchar(12),'14:00:00.000'))"}

 

kwargs.update(query)

 

 

#Putting sql result in a dataframe

r =_run_db_query(**kwargs)

columns = "SRnumber CIF FstName LstName OrgCif OrgName OpenDate ProductLine ProductSubType Transactiontype TransactionSubType Status Administrator Owner".split()

df = pd.DataFrame.from_records(r, columns=columns)

 

 

 

#selecting the excel sheet

sheet1 = wb.get_sheet_by_name('Allocation')

         

 

#Get list of all consultant names

consultants = []

for row_cells in sheet1.iter_rows(min_row=1, max_row=1):

    for cell in row_cells:

        if not cell.value == None:

            consultants.append(cell.value) 

 

#consultants here or not

values = []

for row_cells in sheet1.iter_rows(min_row=2, max_row=2):

    for cell in row_cells:

        if not cell.value == str(cell.value) and cell.value != None:

            values.append(cell.value) 

         

# creating dictionary of all consultants and the value of their productivity as long as productivity is > 0

prod_dict = {}

for x, y in zip(consultants, values):

    if y > 0:

        try:

            prod_dict[x].append(y)

        except:

            prod_dict[x] = []

            prod_dict[x].append(y)

 

def strip_str(value_list):

    return [x.strip() for x in value_list]

 

 

# Joining two columns

Tlist = []

for i,r in df.iterrows():

    if r['TransactionSubType'] is not None:

        r['Type'] = f"{r['Transactiontype']} {r['TransactionSubType']}"

        Tlist.append(r['Type'].strip())

    else:

        r['Type'] = r['Transactiontype']

        Tlist.append(r['Type'].strip())

df['Type'] = Tlist

 

#Identifying Endowments

myEndowmentlist  = []     

for i, r in df.iterrows():

    if r['ProductSubType'] != 'Pure Endowment' :

        r['Type'] = r['Type']

        myEndowmentlist.append(r['Type'])

    else:

        r['Type'] = "Pure Endowment"+' '+r['Type']

        myEndowmentlist.append(r['Type'])

df['Type'] = myEndowmentlist

 

 

# Adding a new column where we shall put the weighting of each instruction in

df['Weight'] = ''

pd.to_numeric(df['Weight'])

 

#creating dictionary of weights for contacts

workbook = xlrd.open_workbook(r"Allocation sheet.xlsx")

sheet = workbook.sheet_by_name('WeightingContact')

col_a = sheet.col_values(0, 1)

col_b = sheet.col_values(1, 1)

ContactW = {a : b for a, b in zip(strip_str(col_a), col_b)}

 

#creating dictionary of weights for Organisations

sheet2 = workbook.sheet_by_name('WeightingORG')

col_x = sheet2.col_values(0, 1)

col_z = sheet2.col_values(1, 1)

OrgW = {a : b for a, b in zip(strip_str(col_x), col_z)}

 

 

#Getting the weighting for the transaction in question

def ContactWeight(Type):        #For contact

    weight = ContactW.get(Type)  

    return weight

 

def OrgWeight(Type):            #For organisation

    weight = OrgW.get(Type)

    return weight

 

#Assigning weights to weight column in DF. The below way of doing it might cause a problem

Wlist   = []     

for i, r in df.iterrows():

    if r['OrgCif'] is None:

        r['Weight'] = ContactWeight(r['Type'])

        Wlist.append(r['Weight'])

    else:

        r['Weight'] = OrgWeight(r['Type'])

        Wlist.append(r['Weight'])

df['Weight'] = Wlist

 

 

#Adding the word legal to the type column if the sr is for a legal

mytypelist  = []     

for i, r in df.iterrows():

    if r['OrgCif'] is None:

        r['Type'] = r['Type']

        mytypelist.append(r['Type'])

    else:

        r['Type'] = "Legal"+' '+r['Type']

        mytypelist.append(r['Type'])

df['Type'] = mytypelist

 

#changing the weighting of all NaN values to 0 in the df

df['Weight'].fillna(0, inplace=True)

 

#Sort the df in descending order by weights. This is important for more accurate allocation

df = df.sort_values(by=['Weight'],ascending=False)

 

 

#Creating a dictionary called elm which contains what every consultant is trained on

workbook = xlrd.open_workbook('Allocation sheet.xlsx')

workbook = xlrd.open_workbook('Allocation sheet.xlsx', on_demand = True)

worksheet = workbook.sheet_by_name('Trained')

first_row = [] # The row where we stock the name of the column

for col in range(worksheet.ncols):

    first_row.append(worksheet.cell_value(0,col).strip() )

 

data =[]

elm = {}

for col in range(worksheet.ncols):

    for row in range(1, worksheet.nrows):

        data.append(worksheet.cell_value(row,col).strip())

    elm[first_row[col]] = data

    data =[]

 

 

#Create a dictionary with all the instructions and their SR numbers

TransactionDict = {}

for i, r in df.iterrows():

    try:

        TransactionDict[r['Type']].append(r['SRnumber'])

    except KeyError:

        TransactionDict[r['Type']] = []

        TransactionDict[r['Type']].append(r['SRnumber'])

       

 

#get a list of consultants that are here and trained on the instruction

def here_and_trained(key):

    consultants_trained = elm.get(key)

    consultants_available = list(prod_dict.keys())

    here_and_trained_list = list(set(consultants_trained) & set(consultants_available))

    return here_and_trained_list

 

 

#checks if json file exists.If it doesnt then an empty dict gets created with the consultants that are here the day as keys

if not os.path.exists(working_directory+'\\JsonFiles\\'+str(today_date)+'.json'):

    import Allocator1 

else:

    import Allocator2

 

 

#######################################################################################################################################

 

 

import os

import openpyxl as pyxl

import json

import datetime

from collections import defaultdict

import random

import pandas as pd

from pandas import ExcelWriter

from collections import Iterable

import pyodbc

 

today_date = datetime.date.today()

today_datetime = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

 

 

#setting path variables

working_directory = 'C:\\Users\\jacoen\\Desktop\\Allocation'

 

 

 

#Set working directory

os.chdir(working_directory)

 

wb = pyxl.load_workbook('Allocation sheet.xlsx')

 

 

from Main import TransactionDict

from Main import here_and_trained

from Main import OrgW

from Main import ContactW

from Main import prod_dict

 

############################################################################################################

    

#deciding algorithm 1. This will kick off if it is the first round allocation.

consultant_allocations = defaultdict(lambda: defaultdict(list))   

def allocation(consultant, Type):

    y = [SR for SR in TransactionDict[Type]] 

    y = y[0]

    consultant_allocations[consultant][Type].append(y)

    TransactionDict[Type].remove(y)  

    update_weights()                                 

    return consultant_allocations          

 

def update_weights():

    getweights()

    getweights2() 

 

def first_round_algoritm(Type):

    results = {}

    consultantlist = []

    for consultant in here_and_trained(Type):

            if consultant not in consultant_allocations.keys():

                consultantlist.append(consultant)

                return random.choice(consultantlist)

            else:

                if not consultantlist:

                        for consultant in here_and_trained(Type):

                            try:

                                results[consultant]=(([totalweight_return(consultant)]+[total_type_weight(consultant, Type)]))

                            except KeyError:

                                results[consultant] = []

                                results[consultant]=(([totalweight_return(consultant)]+[total_type_weight(consultant, Type)]))

                        gets_sr = min(results, key=results.get)      

                        return gets_sr

                       

def themain():

        for Type in TransactionDict.keys():

            try:

                while not TransactionDict.get(Type) == []:

                        consultant = first_round_algoritm(Type)

                        allocation(consultant, Type)

            except IndexError:

                TransactionDict.pop(Type)

                themain()

        with open(working_directory+'\\JsonFiles\\'+str(today_date)+'.json', 'w') as outfile: 

            json.dump((consultant_allocations), outfile, indent=4)

           

        

#########################################################################################################

 

 

########################################################################################################           

 

 

#first create a new dictionary for Orgs that includes the word legal before the transaction type.

OrgW2 = {("Legal "+a) : b for a, b in OrgW.items()}

 

 

#Create a dictionary of weights for every single transaction type and subtype.

combined_weights = {**ContactW, **OrgW2}

 

 

#Getting the weighting for the transaction in question

def totalweight(Type):            #For total

    weight = combined_weights.get(Type)

    return weight

                   

 

#This gets the total number of SRs that the consultant has received thus far.

def total_number_of_srs(consultant):

    count_of_srs = 0

    if consultant in consultant_allocations.keys():

        for Type, SR in consultant_allocations[consultant].items():

            cnt = len(SR)

            count_of_srs = count_of_srs + cnt

        return count_of_srs

    else:

        return 1

     

#This returns the total type weight

def total_type_weight(consultant, Type):

    #print('This is the error type',Type)

    consultant_allocations_weighting()

    if consultant in consultant_allocations_with_weights.keys():

        if Type in consultant_allocations_with_weights[consultant]:

            return consultant_allocations_with_weights[consultant][Type]  

        else: return 0

    else: return 0   

    

    

#creating a dictionary showing the total weighting of SRs allocated to each consultant

consultant_weights = {}

consultant_weights2 = {}

err_list = []

 

 

#this returns the specific consultant's productivity. It's used in the getweights2 function.

def productivity(consultant):

    if consultant in prod_dict.keys():

        productivity = prod_dict.get(consultant)

        return float(productivity[0])

    else:

        return 1

 

 

#this uses getweights func to sum all the values in consultant_weights    

def getweights():

    consultant_weights.clear()

    for x in consultant_allocations.keys():

        try:

            for Type, SR in consultant_allocations[x].items():

                try:

                    consultant_weights[x].append(totalweight(Type)*len(SR))

                except KeyError:

                    consultant_weights[x] = []

                    consultant_weights[x].append(totalweight(Type)*len(SR))

        except TypeError:

            err_list.append(Type)

            print('This is the error list',err_list)

 

              

def getweights2():

    for z,w in consultant_weights.items():

        try:

            consultant_weights2[z] = ((sum(consultant_weights[z]))/productivity(z))

        except KeyError:

            consultant_weights2[z] = []

            consultant_weights2[z] = ((sum(consultant_weights[z]))/productivity(z))

    return consultant_weights2

 

 

#This returns the consultant total weight           

def totalweight_return(consultant):

    if consultant in consultant_weights2.keys():

        return consultant_weights2.get(consultant)

    else:

        return 0

 

 

 

#create a dictionary(consultant_allocations_with_weights2) showing what the total weight is for each SR type that the consultant has received.

#The below is used in the allocator2.

consultant_allocations_with_weights = defaultdict(lambda : defaultdict(list))

def consultant_allocations_weighting():

    for x in consultant_allocations.keys():

            for Type, SR in consultant_allocations[x].items():

                consultant_allocations_with_weights[x][Type] = totalweight(Type)*len(SR)    

    return consultant_allocations_with_weights

                   

 

 

 

themain()

 

consultant_allocations_weighting()

 

 

#This writes the dictionaries to a json file

with open(working_directory+'\\JsonFiles\\{} consultant_allocations_with_weights.json'.format(str(today_date)), 'w') as outfile: 

    json.dump((consultant_allocations_with_weights), outfile, indent=4)

 

with open(working_directory+'\\JsonFiles\\{} consultant_weights.json'.format(str(today_date)), 'w') as outfile: 

    json.dump((consultant_weights), outfile, indent=2)

 

with open(working_directory+'\\JsonFiles\\{} consultant_weights2.json'.format(str(today_date)), 'w') as outfile: 

    json.dump((consultant_weights2), outfile, indent=2)

 

 

#Creating sql update statement

 

 

#Turns list of list into a list

def flatten(sr_list):

    for i in sr_list:

            if isinstance(i, Iterable) and not isinstance(i, str):

                for subc in flatten(i):

                    yield subc

            else:

                yield i

   

#This creates the sql update statement for each consultant

#def individual_sql_query(consultant):

#    sr_list = []

#    for a,b in consultant_allocations[consultant].items():

#        sr_list.append(b)

#    sr_list = flatten(sr_list)

#    _update_query = 'UPDATE S_SRV_REQ SET OWNER_EMP_ID = (SELECT ROW_ID FROM S_USER WHERE LOGIN = '+"'{}')".format(consultant)+' where SR_NUM in (' + ','.join((str("'{}'").format(n) for n in sr_list))+ ')'

#    return _update_query

#

#

#

#def output_complete_sql():

#    sql = ""

#    for consultant in consultant_allocations.keys():

#        sql += ' ' + str(individual_sql_query(consultant))

#    print('this is complete sql',sql)   

#    return sql

#    

 

 

#Below is the sql update statement back to the siebeldb

 

#def _create_db_connection(**kwargs):

#    """

#    Establish a connection to the DB

#    :param kwargs: db_connection(server name or IP)

#    :param kwargs: database

#    :param kwargs: username

#    :param kwargs: password

#    :return: The DB connection object

#    """

#    driver_type = '{ODBC Driver 13 for SQL Server}'

#    connection_str = f'DRIVER={driver_type};' \

#                     f'SERVER={kwargs.get("db_connection")};' \

#                     f'DATABASE={kwargs.get("database")};UID={kwargs.get("username")};' \

#                     f'PWD={kwargs.get("password")}'

#    try:

#        connection = pyodbc.connect(connection_str)

#    except pyodbc.OperationalError as error_msg:

#        msg = f'Verify that the server value is correct: {kwargs.get("db_connection")}'

#        raise Exception(f'{msg}\n{error_msg.args}')

#    return connection

#

#kwargs = {'db_connection': 'sql-siebel.int.gray.net,64495',

#                           'username': 'sadmin',

#                          'password': 'sadminint',

#                          'database': 'siebeldb'}

#

#def _run_db_update_query(**kwargs):

#    """

#    Run any query

#    :param kwargs: db_connection

#    :param kwargs: database

#    :param kwargs: username

#    :param kwargs: password

#    :param kwargs: query

#    :return: a tuple of the query results connection object

#    """

#   

#    connection =  _create_db_connection(**kwargs)

#

#    with connection:

#        cursor = connection.cursor()

#    #query = output_complete_sql()

#    print('jhnbdhbagfh')

#    return cursor.execute(output_complete_sql()), connection.commit()

#

#_run_db_update_query(**kwargs)

 

 

 

 

 

#Need to visualize the results below

visualize = pd.DataFrame.from_dict(consultant_allocations_with_weights)

cols_headers = visualize.columns.values.tolist()

visualize.loc['Total Weighting'] = visualize.sum()

 

 

writer = ExcelWriter(working_directory+'\\AllocationResults\\Allocation_result.xlsx')

 

visualize.to_excel(writer,'Sheet1',index=True)

writer.save()

 

#send out the email with the allocation result attached

import win32com.client

outlook = win32com.client.Dispatch("Outlook.Application")

mail = outlook.CreateItem(0)

mail.To = "jaco.enslin@allangray.co.za" 

mail.subject = '{} allocation result'.format(str(today_datetime))

mail.Body = "Allocation result is attached"

allocation_result = working_directory+'\\AllocationResults\\Allocation_result.xlsx'

mail.Attachments.Add(allocation_result)

mail.Send()

 

 

###########################################################################################################################################

 

 

 

import os

import openpyxl as pyxl

import json

import datetime

from collections import defaultdict

import pandas as pd

from pandas import ExcelWriter

from collections import Iterable

import pyodbc

 

 

#setting path variables

working_directory = 'C:\\Users\\jacoen\\Desktop\\Allocation'

 

#Set working directory

os.chdir(working_directory)

 

wb = pyxl.load_workbook('Allocation sheet.xlsx')